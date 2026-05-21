from openpyxl import *

# Criar uma planilha(book)
book = Workbook()
# Como visualizar páginas existentes
print(book.sheetnames)
# Como criar uma página
book.create_sheet('Frutas')
# Como selecionar uma página
frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Fruta 2', '2', 'R$15,90'])
frutas_page.append(['Fruta 3', '10', 'R$30,90'])
frutas_page.append(['Fruta 4', '2', 'R$50,50'])
# Salvar a planilha
book.save(r'C:\Users\chris\Desktop\scripts-python\manipular-planilha\Planilha de compras.xlsx')
