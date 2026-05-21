from openpyxl import *

# Carregando arquivo
book = load_workbook(r'C:\Users\chris\Desktop\scripts-python\manipular-planilha\Planilha de compras.xlsx')
# Selecionando uma página
frutas_page = book['Frutas']
# Imprimindos os dados de cada linha
for rows in frutas_page.iter_rows(min_row=2,max_row=5):
    print(f'{rows[0].value},{rows[1].value},{rows[2].value}')
# Alterar informação da célula
for rows in frutas_page.iter_rows(min_row=2,max_row=5):
    for cell in rows:
        if cell.value == 'Banana':
            cell.value = 'Fruta 1'

# Salvar as alterações em uma nova planilha
book.save(r'C:\Users\chris\Desktop\scripts-python\manipular-planilha\Planilha de compras v2.xlsx')