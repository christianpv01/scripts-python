'''
# DESAFIO <Dev Aprender>

## Crie uma planilha com os seguintes dados e depois poste nos comentários o código que usou para criar essa planilha

# Nome da Planilha
meus computadores

# Nome da página
computadores

# Nome das colunas
Eletrônica, memória ram, preço

# Dados
Computador 1, 8gb Ram, R$2500
Computador 2, 16gb Ram, R$5500
Computador 3, 32gb Ram, R$8500
'''
from openpyxl import *

# Criando a planilha
book = Workbook()
# Criando a aba
book.create_sheet('computadores')
# Selecionando a página
computadores = book['computadores']
# Nomeando as colunas e adicionando os dados
computadores.append(['Eletrônica','Memória RAM','Preço'])
computadores.append(['Computador 1','8 GB','R$ 2.500'])
computadores.append(['Computador 2','16 GB','R$ 5.500'])
computadores.append(['Computador 3','32 GB','R$ 8.500'])
# Salvando a planilha
book.save(r'C:\Users\chris\Desktop\scripts-python\manipular-planilha\meus computadores.xlsx')
# Carregando a planilha
book = load_workbook(r'C:\Users\chris\Desktop\scripts-python\manipular-planilha\meus computadores.xlsx')
# Selecionando a página
computadores = book['computadores']
# Lendo a planilha
for rows in computadores.iter_rows():
    print(f'|{rows[0].value:^20}|{rows[1].value:^11}|{rows[2].value:^10}|')